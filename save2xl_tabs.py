#!/usr/bin/python -tt
# Project: netmiko38
# Filename: save2xl_tabs
# claudia
# PyCharm

from __future__ import absolute_import, division, print_function

__author__ = "Claudia de Luna (claudia@indigowire.net)"
__version__ = ": 1.0 $"
__date__ = "10/13/20"
__copyright__ = "Copyright (c) 2018 Claudia"
__license__ = "Python"

import argparse
import openpyxl
import datetime
import os

def create_xlwkbk(debug=False):

    if debug: print(f"\t--- Creating an Excel Workbook Object.\n")
    workbook = openpyxl.Workbook()

    return workbook


def create_xltab(xlobj, xltab_name, debug=False):

    if debug: print(f"\t--- Creating Tab {xltab_name} in Excel file. \n")
    current_tabs = xlobj.sheetnames

    if xltab_name in current_tabs:
        print(f"\tTab already exists!\n")
    else:
        xlobj.create_sheet(xltab_name)


def open_xlwkbk(xlfilepath, debug=False):

    if debug: print(f"\t--- Open existing Excel Workbook {xlfilepath}.\n")
    workbook = openpyxl.load_workbook(filename=xlfilepath)

    return workbook


def save2tab(xlobj, xltab_name, data_list, debug=False):
    """
    Function to save data to the provided tab
    Funciton checks to see if the tab already exists and if it does not it creates it.
    :param xl0bj:  Excel Workbook Object
    :param xltab_name:  Excel Workbook tab or sheet to save data to
    :param data_list:  Data to save to specified tab or sheet.  This should be a list of lists
    :return:
    """
    if debug: print(f"\t--- Saving Data to Tab {xltab_name}.\n")
    if xltab_name not in xlobj.sheetnames:
        create_xltab(xlobj, xltab_name)

    ws = xlobj[xltab_name]

    # The tricky part is manipulating the data into the cells
    # In this block first we check to see if "data_list" is actually a list
    if type(data_list) == list:
        # then we have to check to see if what is in each row is a list or a string
        for row in data_list:
            if type(row) == list:
                # the append method expects an iterable (list, tuple)
                ws.append(row)
            elif type(row) == str:
                # if its a string then you want to cast is as a list otherwise you will get a letter in each cell
                ws.append([row])
    # if data-list is actually a string we just put the string in the first cell.
    # This is not what you want so you will be in the business of manipulating your text as above
    elif type(data_list) == str:
        ws['A1'] = data_list


def save_xlwkbk(xlobj, xlfilename, xlpath=os.getcwd(), debug=False):

    if debug: print(f"\t--- Saving the Workbook Object as {xlfilename} to path {xlpath}.\n")
    fp = os.path.join(xlpath, xlfilename)
    result = xlobj.save(fp)

    return result, fp


def get_show_output():
    print(f"\t--- Dummy function to load test show command data.\n")
    response = """
    Smart Licensing Status: Smart Licensing is DISABLED

    cisco CSR1000V (VXE) processor (revision VXE) with 2392579K/3075K bytes of memory.
    Processor board ID 96NBTE4ZXYT
    3 Gigabit Ethernet interfaces
    32768K bytes of non-volatile configuration memory.
    8113280K bytes of physical memory.
    7774207K bytes of virtual hard disk at bootflash:.
    0K bytes of WebUI ODM Files at webui:.

    Configuration register is 0x2102

        """
    return response


def main():

    """
    Example script for openpyxl and saving show command output to an Excel file including different tabs/sheets.

    Virtual Environment Requirements: requirements-netmiko-pandas-xl.txt
    Returns:

    """

    datestamp = datetime.date.today()
    print(f"===== Date is {datestamp} ====")

    # Define an Excel Workbook filename based on the name passed as an argument in the command line
    # python save2xl_tabs.py -f CLOD
    # the CLI command above creates a file CLOD_YYYY-MM-DD.xlsx
    # or using the default file name DEFAULT
    xlfilename = f"{arguments.filename}_{datestamp}.xlsx"

    # Note on the functions. Many have an optional parameter "debug" that just prints information when set to True
    # It is set to False by default but sometimes its handy to know where you are in the script

    # Create an Excl Workbook Object
    xobj = create_xlwkbk(debug=True)

    # Add a Tab to the Excel Workbook Ojbect
    create_xltab(xobj, "TEST", debug=True)

    # Save data to a Tab
    # Sample output string
    output = get_show_output()

    # Take the string response and split by lines
    list_of_output = output.splitlines()

    # Send the raw string to be saved to a tab called RAW Output
    save2tab(xobj,"RAW Output", output, debug=True)
    # Send the list of lines to be saved to a tab called LIST Output
    save2tab(xobj,"LIST Output", list_of_output, debug=True)

    # Save the Excel Workbook Object to Disk
    res, filepath = save_xlwkbk(xobj, xlfilename, debug=True)

    # Open the saved Workbook back up again and display the Tabs
    wkbk = open_xlwkbk(filepath)
    print(f"Saved Excel file {filepath} \n  has the following Tabs: \n\t{wkbk.sheetnames}\n\n")


# Standard call to the main() function.
if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Script Description",
                                     epilog="Usage: ' python save2xl_tabs' ")

    parser.add_argument('-f', '--filename', help='Excel filename. Default=DEFAULT-YYYY-MM_DD', action='store',
                        default="DEFAULT")
    arguments = parser.parse_args()
    main()
